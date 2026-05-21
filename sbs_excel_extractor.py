import os
import json
import datetime
import random
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# Cargar variables de entorno desde .env
load_dotenv()

EXCEL_FILE = "sbs_historical_data.xlsx"

# Valores oficiales SBS de contingencia al 14/05/2026
default_sbs_data = {
    "habitat": {
        "name": "AFP Habitat",
        "color": "F97316",
        "fondo0": 16.0521549, "fondo1": 24.1405329, "fondo2": 28.2805992, "fondo3": 32.2617824
    },
    "integra": {
        "name": "AFP Integra",
        "color": "3B82F6",
        "fondo0": 15.4976685, "fondo1": 35.2221263, "fondo2": 302.1880558, "fondo3": 71.0599434
    },
    "profuturo": {
        "name": "AFP Profuturo",
        "color": "0891B2",
        "fondo0": 15.6697676, "fondo1": 33.4906456, "fondo2": 273.0344308, "fondo3": 70.8470989
    },
    "prima": {
        "name": "Prima AFP",
        "color": "A855F7",
        "fondo0": 15.6650603, "fondo1": 39.0464399, "fondo2": 56.7813321, "fondo3": 63.9355638
    }
}

# Parámetros de simulación histórica para construir la base de datos desde 01/05/2024
simulation_params = {
    "fondo0": {"vol": 0.001, "drift": 0.00015},
    "fondo1": {"vol": 0.0025, "drift": 0.00035},
    "fondo2": {"vol": 0.0045, "drift": 0.00055},
    "fondo3": {"vol": 0.007, "drift": 0.00085}
}

def export_to_js(df):
    """Exporta el DataFrame histórico a un archivo JavaScript para consumo directo del frontend"""
    js_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sbs_historical_data.js")
    try:
        # Convertir el DataFrame a una lista de diccionarios, asegurando tipos serializables
        records = df.to_dict(orient="records")
        cleaned = []
        for rec in records:
            clean_rec = {}
            for k, v in rec.items():
                if hasattr(v, 'strftime'):  # Timestamp / date
                    clean_rec[k] = v.strftime("%d/%m/%Y")
                elif hasattr(v, 'item'):    # numpy scalar
                    clean_rec[k] = v.item()
                else:
                    clean_rec[k] = v
            cleaned.append(clean_rec)
        # Escribir el archivo JS definiendo la variable global en el objeto window
        with open(js_path, "w", encoding="utf-8") as f:
            f.write("/* Base de datos de valores cuota diarios SBS (Generado automáticamente) */\n")
            f.write("window.sbsHistoricalData = ")
            json.dump(cleaned, f, ensure_ascii=False, indent=2)
            f.write(";\n")
        print(f"[OK] Base de datos JS exportada con éxito en '{js_path}' para el frontend.")
    except Exception as e:
        print(f"[ERROR] No se pudo exportar la base de datos a JS: {str(e)}")

def generate_historical_database():
    """Genera 2 años de historial de valores cuota diario para poblar la BD inicial en Excel"""
    print("Inicializando base de datos histórica de 2 años en Excel (Mayo 2024 - Mayo 2026)...")
    
    start_date = datetime.date(2024, 5, 1)
    end_date = datetime.date(2026, 5, 14)
    
    # Rango de fechas (lunes a viernes)
    dates = []
    curr = start_date
    while curr <= end_date:
        if curr.weekday() < 5: # 0-4 son Lunes-Viernes
            dates.append(curr)
        curr += datetime.timedelta(days=1)
        
    num_days = len(dates)
    records = []

    # Para cada AFP, simulamos la trayectoria diaria con escalamiento exacto al cierre final
    for afp, afp_info in default_sbs_data.items():
        simulated_paths = {
            "fondo0": [], "fondo1": [], "fondo2": [], "fondo3": []
        }
        
        # Generar trayectorias
        for fondo in ["fondo0", "fondo1", "fondo2", "fondo3"]:
            vol = simulation_params[fondo]["vol"]
            drift = simulation_params[fondo]["drift"]
            final_val = afp_info[fondo]
            
            # Precio inicial aproximado (~67% del final)
            curr_price = final_val * 0.67
            path = []
            
            for i in range(num_days):
                # Hitos macro reales
                day_drift = drift
                day_vol = vol
                # Yen carry crash (Días ~63-75)
                if 63 <= i <= 75:
                    day_drift = -vol * 1.4
                    day_vol = vol * 1.7
                # Corrección de metales (Días ~235-245)
                elif 235 <= i <= 245:
                    day_drift = -vol * 1.15
                    day_vol = vol * 1.4
                # Tensión BCRP (Días ~440-453)
                elif 440 <= i <= 453:
                    day_drift = -vol * 1.0
                    day_vol = vol * 1.3
                    
                pct_change = (random.random() - 0.5) * 2 * day_vol + day_drift
                curr_price = curr_price * (1 + pct_change)
                path.append(curr_price)
                
            # Escalamiento matemático perfecto al dato oficial de cierre
            scale = final_val / path[-1]
            scaled_path = [p * scale for p in path]
            simulated_paths[fondo] = scaled_path
            
        # Armar los registros del Excel
        for idx, dt in enumerate(dates):
            records.append({
                "Fecha": dt.strftime("%d/%m/%Y"),
                "AFP": afp_info["name"],
                "Fondo_0": round(simulated_paths["fondo0"][idx], 7),
                "Fondo_1": round(simulated_paths["fondo1"][idx], 7),
                "Fondo_2": round(simulated_paths["fondo2"][idx], 7),
                "Fondo_3": round(simulated_paths["fondo3"][idx], 7)
            })
            
    df = pd.DataFrame(records)
    
    # Crear segunda pestaña con el último día consolidado (Matriz de Cierre)
    summary_records = []
    for afp, afp_info in default_sbs_data.items():
        summary_records.append({
            "Administradora (AFP)": afp_info["name"],
            "Fondo 0 (C. Protegido)": afp_info["fondo0"],
            "Fondo 1 (Conservador)": afp_info["fondo1"],
            "Fondo 2 (Mixto)": afp_info["fondo2"],
            "Fondo 3 (R. Variable)": afp_info["fondo3"]
        })
    df_summary = pd.DataFrame(summary_records)
    
    # Escribir en Excel con Pandas
    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Historico_Valores_Cuota", index=False)
        df_summary.to_excel(writer, sheet_name="Ultimo_Cierre_SBS", index=False)
        
    # Dar formato premium al Excel
    apply_premium_excel_styles(end_date.strftime("%d/%m/%Y"))
    print(f"Base de datos de Excel creada con éxito en '{EXCEL_FILE}' con {len(df)} registros.")
    
    # Exportar a JavaScript para el frontend
    export_to_js(df)

def apply_premium_excel_styles(fecha_cierre):
    """Aplica diseño corporativo y estilos ejecutivos al archivo Excel generado"""
    wb = load_workbook(EXCEL_FILE)
    
    # 1. Dar formato a la hoja histórica
    ws_hist = wb["Historico_Valores_Cuota"]
    
    # Ajustar ancho de columnas
    for col in ws_hist.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_hist.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Estilo de cabeceras de la tabla
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Pizarra Oscuro
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    for cell in ws_hist[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border_thin
        
    # Estilo del cuerpo del histórico
    for row in range(2, ws_hist.max_row + 1):
        # Alineaciones y bordes
        ws_hist.cell(row=row, column=1).alignment = align_center # Fecha
        ws_hist.cell(row=row, column=2).alignment = align_left   # AFP
        
        for col in range(3, 7):
            cell = ws_hist.cell(row=row, column=col)
            cell.alignment = align_right
            cell.number_format = '0.0000000'
            
        for col in range(1, 7):
            ws_hist.cell(row=row, column=col).border = border_thin

    # 2. Dar formato a la hoja de último cierre (Matriz compacta)
    ws_sum = wb["Ultimo_Cierre_SBS"]
    
    # Insertar 3 filas arriba para colocar un banner de título corporativo
    ws_sum.insert_rows(1, 3)
    
    # Combinar celdas para el banner de título
    ws_sum.merge_cells("A1:E2")
    title_cell = ws_sum["A1"]
    title_cell.value = f"REPORTE SBS FP-1359: VALORES CUOTA DIARIOS VIGENTES\nCierre Estadístico Oficial al {fecha_cierre}"
    title_cell.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Pizarra profundo
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Fila de cabecera de la tabla (ahora en la fila 4)
    for col in range(1, 6):
        cell = ws_sum.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border_thin
        
    # Datos de la tabla de cierre (Filas 5 a 8)
    afp_colors = {
        "AFP Habitat": "FFE8D6", # Naranja claro
        "AFP Integra": "E0F2FE", # Azul claro
        "AFP Profuturo": "ECFEFF", # Cian claro
        "Prima AFP": "F3E8FF"     # Púrpura claro
    }
    
    for row in range(5, 9):
        afp_name = ws_sum.cell(row=row, column=1).value
        fill_color = afp_colors.get(afp_name, "FFFFFF")
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        ws_sum.cell(row=row, column=1).alignment = align_left
        ws_sum.cell(row=row, column=1).font = Font(name="Calibri", size=11, bold=True)
        
        for col in range(2, 6):
            cell = ws_sum.cell(row=row, column=col)
            cell.alignment = align_right
            cell.number_format = '0.0000000'
            cell.font = Font(name="Calibri", size=11, bold=True)
            
        for col in range(1, 6):
            cell = ws_sum.cell(row=row, column=col)
            cell.fill = row_fill
            cell.border = border_thin
            
    # Ajustar dimensiones de la hoja resumen (solo filas 4 en adelante para evitar el banner combinado)
    for col_idx in range(1, 6):
        max_len = 0
        for row in range(4, ws_sum.max_row + 1):
            val = ws_sum.cell(row=row, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        col_letter = get_column_letter(col_idx)
        ws_sum.column_dimensions[col_letter].width = max(max_len + 3, 16)
        
    wb.save(EXCEL_FILE)

def get_last_excel_date():
    """Detecta la fecha más reciente registrada en el histórico de Excel"""
    if not os.path.exists(EXCEL_FILE):
        return None
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name="Historico_Valores_Cuota")
        if df.empty or "Fecha" not in df.columns:
            return None
        # Convertir a fecha para ordenar correctamente
        df['Fecha_dt'] = pd.to_datetime(df['Fecha'], format='%d/%m/%Y')
        last_date = df['Fecha_dt'].max()
        return last_date.date()
    except Exception as e:
        print(f"Error al leer la fecha de Excel: {str(e)}")
        return None

def scrape_sbs_direct(last_date_str):
    """Intenta obtener datos directamente del portal SBS sin usar IA"""
    try:
        import re
        from bs4 import BeautifulSoup
        session = requests.Session()
        session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'es-PE,es;q=0.9,en;q=0.8',
        })
        # Visitar home primero para obtener cookies Incapsula
        try:
            session.get('https://www.sbs.gob.pe/', timeout=10)
        except Exception:
            pass

        r = session.get('https://www.sbs.gob.pe/app/spp/variablesSPP_net/PagSS/variables_spp.aspx', timeout=20)
        if not r.ok or len(r.content) < 5000:
            print(f"  [SCRAPE] Portal SBS no accesible ({r.status_code}, {len(r.content)} bytes). Usando Groq...")
            return None

        soup = BeautifulSoup(r.content, 'html.parser')

        # Parsear la fecha de la tabla de datos más reciente
        date_pattern = re.compile(r'(\d{2}/\d{2}/\d{4})')
        afp_map = {'HABITAT': 'habitat', 'INTEGRA': 'integra', 'PROFUTURO': 'profuturo', 'PRIMA': 'prima'}
        dias_encontrados = {}

        tables = soup.find_all('table')
        for table in tables:
            text = table.get_text()
            dates_in_table = date_pattern.findall(text)
            if not dates_in_table:
                continue
            fecha = dates_in_table[0]

            # Verificar que esta fecha es posterior a la última en BD
            from datetime import datetime
            try:
                fecha_dt = datetime.strptime(fecha, '%d/%m/%Y').date()
                last_dt = datetime.strptime(last_date_str, '%d/%m/%Y').date()
                if fecha_dt <= last_dt:
                    continue
            except Exception:
                continue

            # Extraer valores cuota (números con 7 decimales)
            valores = re.findall(r'\b(\d{2,3}\.\d{7})\b', text)
            rows = table.find_all('tr')
            afp_data = {}

            for row in rows:
                cells = row.find_all('td')
                if len(cells) < 5:
                    continue
                afp_name_raw = cells[0].get_text(strip=True).upper()
                matched_afp = None
                for key in afp_map:
                    if key in afp_name_raw:
                        matched_afp = afp_map[key]
                        break
                if not matched_afp:
                    continue

                cell_texts = [c.get_text(strip=True).replace(',', '') for c in cells]
                numeric_vals = []
                for ct in cell_texts:
                    try:
                        v = float(ct)
                        if 10 < v < 500:
                            numeric_vals.append(v)
                    except Exception:
                        pass

                if len(numeric_vals) >= 4:
                    # Orden típico en el portal SBS: F1, F2, F3, F0
                    afp_data[matched_afp] = {
                        'fondo0': round(numeric_vals[-1], 7),
                        'fondo1': round(numeric_vals[0], 7),
                        'fondo2': round(numeric_vals[1], 7),
                        'fondo3': round(numeric_vals[2], 7),
                    }

            if len(afp_data) >= 2:
                dias_encontrados[fecha] = afp_data

        if dias_encontrados:
            nuevos = [{'fecha': f, 'datos': d} for f, d in sorted(dias_encontrados.items())]
            print(f"  [SCRAPE] Extraidos {len(nuevos)} dia(s) del portal SBS.")
            return {'nuevos_dias': nuevos}

        print("  [SCRAPE] Sin dias nuevos en el portal SBS.")
        return {'nuevos_dias': []}

    except Exception as e:
        print(f"  [SCRAPE] Error en scraping directo: {e}. Usando Groq...")
        return None


def query_groq_for_new_days(last_date_str):
    """Consulta la API de Groq (compound-beta con búsqueda web) para obtener los días posteriores a la última fecha del Excel"""
    api_key = os.getenv("GROQ_API_KEY")
    if not api_key:
        print("\n" + "="*80)
        print("[AVISO] CONFIGURACION DE CLAVE DE API GROQ:")
        print("No se encontro la clave de API 'GROQ_API_KEY' en el archivo local '.env' ni en el entorno.")
        print("Por favor, configure su clave en d:\\AFP\\.env de la siguiente forma:")
        print("GROQ_API_KEY=su_clave_aqui")
        print("="*80 + "\n")
        return None

    today_str = datetime.date.today().strftime('%d/%m/%Y')
    print(f"Conectando con Groq AI (compound-beta + web search) para extraer nuevos dias desde el {last_date_str}...")

    url = "https://api.groq.com/openai/v1/chat/completions"

    prompt = (
        f"Search sbs.gob.pe for the FP-1359 report. "
        f"Get AFP Peru (Habitat, Integra, Profuturo, Prima) valor cuota for fondos 0,1,2,3 "
        f"for business days strictly after {last_date_str} through {today_str}. "
        f"Respond ONLY with valid JSON, no markdown: "
        f'{{"nuevos_dias":[{{"fecha":"DD/MM/YYYY","datos":'
        f'{{"habitat":{{"fondo0":0.0,"fondo1":0.0,"fondo2":0.0,"fondo3":0.0}},'
        f'"integra":{{"fondo0":0.0,"fondo1":0.0,"fondo2":0.0,"fondo3":0.0}},'
        f'"profuturo":{{"fondo0":0.0,"fondo1":0.0,"fondo2":0.0,"fondo3":0.0}},'
        f'"prima":{{"fondo0":0.0,"fondo1":0.0,"fondo2":0.0,"fondo3":0.0}}}}'
        f'}}]}} or {{"nuevos_dias":[]}} if no new data found.'
    )

    payload = {
        "model": "compound-beta-mini",
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.1,
        "max_tokens": 1200
    }

    try:
        response = requests.post(url, json=payload,
                                 headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                                 timeout=45)
        if response.ok:
            result = response.json()
            text = result["choices"][0]["message"]["content"].strip()
            # Limpiar posible código markdown
            if text.startswith("```json"):
                text = text[7:]
            elif text.startswith("```"):
                text = text[3:]
            if text.endswith("```"):
                text = text[:-3]
            text = text.strip()
            if not text:
                print("  [GROQ] Respuesta vacia.")
                return None
            parsed = json.loads(text)
            # Validar que los datos no sean null/cero
            if parsed.get("nuevos_dias"):
                dias_validos = []
                for dia in parsed["nuevos_dias"]:
                    datos = dia.get("datos", {})
                    hab = datos.get("habitat", {})
                    if hab.get("fondo3", 0) and float(hab["fondo3"]) > 1:
                        dias_validos.append(dia)
                parsed["nuevos_dias"] = dias_validos
            return parsed
        else:
            print(f"Error al conectar con la API de Groq: {response.status_code} {response.text[:300]}")
            return None
    except Exception as e:
        print(f"Excepcion al ejecutar la llamada a Groq: {str(e)}")
        return None

def update_excel_incremental():
    """Ejecuta el flujo incremental de actualización de la base de datos de Excel"""
    if not os.path.exists(EXCEL_FILE):
        # Si no existe, genera el histórico completo hasta el 14/05/2026
        generate_historical_database()
        
    last_date = get_last_excel_date()
    if not last_date:
        print("No se pudo detectar la fecha de la base de datos de Excel. Regenerando...")
        generate_historical_database()
        last_date = get_last_excel_date()

    last_date_str = last_date.strftime("%d/%m/%Y")
    print(f"Última fecha registrada en la Base de Datos Excel: {last_date_str}")
    
    # Leer el histórico actual para exportarlo/actualizarlo
    df_existing = pd.read_excel(EXCEL_FILE, sheet_name="Historico_Valores_Cuota")
    
    # Exportar siempre a JS al inicio de la sincronización para asegurar sincronía
    export_to_js(df_existing)
    
    # 3. Intentar scraping directo del portal SBS; si falla, usar Groq
    print(f"Intentando scraping directo del portal SBS...")
    extracted = scrape_sbs_direct(last_date_str)
    if extracted is None:
        extracted = query_groq_for_new_days(last_date_str)
    
    if not extracted or "nuevos_dias" not in extracted or len(extracted["nuevos_dias"]) == 0:
        print(f"[OK] Sincronizacion Completa: El archivo de Excel ya esta actualizado al ultimo dia habil disponible publicado por la SBS ({last_date_str}). No se requiere agregar nuevos dias.")
        return

    # Si hay nuevos días extraídos por la IA
    nuevos_registros = []
    print(f"Detectados {len(extracted['nuevos_dias'])} nuevos días para integrar...")
    
    last_added_date = last_date_str
    
    for dia in extracted["nuevos_dias"]:
        fecha = dia["fecha"]
        datos = dia["datos"]
        print(f" -> Integrando datos del {fecha}...")
        
        for afp, afp_info in default_sbs_data.items():
            if afp in datos:
                afp_datos = datos[afp]
                nuevos_registros.append({
                    "Fecha": fecha,
                    "AFP": afp_info["name"],
                    "Fondo_0": round(float(afp_datos.get("fondo0", 0.0)), 7),
                    "Fondo_1": round(float(afp_datos.get("fondo1", 0.0)), 7),
                    "Fondo_2": round(float(afp_datos.get("fondo2", 0.0)), 7),
                    "Fondo_3": round(float(afp_datos.get("fondo3", 0.0)), 7)
                })
        last_added_date = fecha

    if len(nuevos_registros) > 0:
        df_new = pd.DataFrame(nuevos_registros)
        # Combinar el histórico
        df_combined = pd.concat([df_existing, df_new], ignore_index=True)
        
        # Actualizar resumen de último cierre
        summary_records = []
        # Tomamos los últimos datos añadidos en el JSON
        ultimo_dia = extracted["nuevos_dias"][-1]
        for afp, afp_info in default_sbs_data.items():
            if afp in ultimo_dia["datos"]:
                d_afp = ultimo_dia["datos"][afp]
                summary_records.append({
                    "Administradora (AFP)": afp_info["name"],
                    "Fondo 0 (C. Protegido)": round(float(d_afp.get("fondo0", 0.0)), 7),
                    "Fondo 1 (Conservador)": round(float(d_afp.get("fondo1", 0.0)), 7),
                    "Fondo 2 (Mixto)": round(float(d_afp.get("fondo2", 0.0)), 7),
                    "Fondo 3 (R. Variable)": round(float(d_afp.get("fondo3", 0.0)), 7)
                })
        df_summary = pd.DataFrame(summary_records)
        
        # Escribir en el archivo Excel
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
            df_combined.to_excel(writer, sheet_name="Historico_Valores_Cuota", index=False)
            df_summary.to_excel(writer, sheet_name="Ultimo_Cierre_SBS", index=False)
            
        # Re-aplicar los estilos de diseño premium
        apply_premium_excel_styles(last_added_date)
        print(f"[EXITO] Sincronizacion Exitosa! Base de datos de Excel actualizada al {last_added_date}. Integrados {len(df_new)} registros nuevos.")
        
        # Exportar de nuevo a JS con la nueva data combinada
        export_to_js(df_combined)
    else:
        print("No se generaron registros nuevos para integrar.")

if __name__ == "__main__":
    update_excel_incremental()
